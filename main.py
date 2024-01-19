import pandas as pd
import random
from openpyxl import workbook
from openpyxl import load_workbook

input_file_path = 'input_output.xlsx'
file_path = 'generated_output.xlsx'
output_file_path = 'NOMAC_Work Order Manangement.xlsx'

final_script = []
script=[]

num_rows_to_read = 1


input_excel_df = pd.read_excel(input_file_path,sheet_name="Sheet1")

names = input_excel_df.iloc[:,0]
lastnames = input_excel_df.iloc[:,1]
comments = input_excel_df.iloc[:,2]
files = input_excel_df.iloc[:,3]
'''
names = ["Johnny","","name3","name4","name5"]
lastnames = ["Abou Haidar","","lname3","lname4","lname5"]
comments = ["sample comment1","test","hello",""]
files='''

#print(random.choice(names))
#print(random.choice(lastnames))

final_result = []

for i in range (100):
    final_result.append([random.choice(names),random.choice(lastnames),random.choice(comments),random.choice(files)])


print(final_result)
final_df=pd.DataFrame(final_result,columns=["first","Lname","comment","file2upload"])
print(final_df)

#final_df.to_excel(file_path,index=False,sheet_name="Sheet1")




final_df.to_excel(file_path,index=False)

types_names = pd.read_excel(input_file_path,sheet_name="Sheet2", nrows=num_rows_to_read).columns.tolist()
Field_names = pd.read_excel(file_path,sheet_name="Sheet1", nrows=num_rows_to_read).columns.tolist()

result_dict = dict(zip(types_names, Field_names))

print(result_dict)

for res in result_dict.keys():
    
    if "File" in res:
        script.append("Upload %{0}% in {1} Field".format(result_dict[res],result_dict[res]))
    else:
        script.append("Enter %{0}% in {1} Field".format(result_dict[res],result_dict[res]))
    




df = pd.read_excel(file_path)

for index, row in df.iterrows():
    singleTC=[]
    for case in script:
        
        for ph in Field_names:
            
            if ph in case:
                singleTC.append(case.replace("%{0}%".format(ph),str(row[ph])))   
                 
    
    final_script.append(singleTC)

wb = load_workbook(output_file_path)
sheets = wb.active
Sheet1 = wb["FSM_NOMAC"]
column =5
new_row=11
script_index=1
for singlef_script in final_script:
    #print(singlef_script)
    
    step_number=1
    Sheet1.cell(column=column-2, row=new_row, value="NOMAC_Process1_TC{0}".format(script_index))
    for i, value in enumerate(singlef_script):

        Sheet1.cell(column=column, row=i+new_row, value=value)
        Sheet1.cell(column=column-1, row=i+new_row, value=step_number)
        Sheet1.cell(column=column+1, row=i+new_row, value="No validation Errors when entered")
        finalrow=i+new_row
        step_number=step_number+1
    
    new_row = finalrow+1
    script_index=script_index+1

wb.save(output_file_path)
